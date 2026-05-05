import serial
import math
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from collections import deque
import queue
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import numpy as np

# --- Configuration ---
PORT = 'COM7'
BAUD_RATE = 115200
MAX_POINTS = 500
SENSOR_IDS = [21, 22, 23]
LOG_DIRECTORY = '.'
CALIBRATION_TIME_SEC = 10.0

start_time = None
is_calibrating = True
calibration_data = {
    'roll': {sid: [] for sid in SENSOR_IDS},
    'pitch': {sid: [] for sid in SENSOR_IDS},
    'yaw': {sid: [] for sid in SENSOR_IDS}
}
baseline_offsets = {
    'roll': {sid: 0.0 for sid in SENSOR_IDS},
    'pitch': {sid: 0.0 for sid in SENSOR_IDS},
    'yaw': {sid: 0.0 for sid in SENSOR_IDS}
}

# --- Initialize Data Storage ---
data = {
    'roll': {sid: deque([0.0] * MAX_POINTS, maxlen=MAX_POINTS) for sid in SENSOR_IDS},
    'pitch': {sid: deque([0.0] * MAX_POINTS, maxlen=MAX_POINTS) for sid in SENSOR_IDS},
    'yaw': {sid: deque([0.0] * MAX_POINTS, maxlen=MAX_POINTS) for sid in SENSOR_IDS}
}

log_queue = queue.Queue()
running = True
is_recording = False
is_marking = False
reset_to_zero = set()

# --- Serial Setup ---
try:
    ser = serial.Serial(PORT, BAUD_RATE, timeout=0.1)
    print(f"Successfully connected to {PORT} at {BAUD_RATE} baud.")
except Exception as e:
    print(f"Error opening serial port: {e}")
    ser = None

# --- Plot Setup ---
fig, (ax_roll, ax_pitch, ax_yaw) = plt.subplots(3, 1, figsize=(10, 8))
fig.canvas.manager.set_window_title('Real-Time Sensor Orientation (Close Look)')
fig.suptitle("Dynamic ±10° Kinematics\nPress SPACE to start/stop recording", fontweight='bold')

lines = {'roll': {}, 'pitch': {}, 'yaw': {}}
colors = {21: 'red', 22: 'green', 23: 'blue'}
axes = [ax_roll, ax_pitch, ax_yaw]
data_keys = ['roll', 'pitch', 'yaw']

for sid in SENSOR_IDS:
    lines['roll'][sid], = ax_roll.plot([], [], color=colors[sid], label=f'Sensor {sid}')
    lines['pitch'][sid], = ax_pitch.plot([], [], color=colors[sid], label=f'Sensor {sid}')
    lines['yaw'][sid], = ax_yaw.plot([], [], color=colors[sid], label=f'Sensor {sid}')

for ax, title in zip(axes, ['Roll', 'Pitch', 'Yaw']):
    ax.set_xlim(0, MAX_POINTS)
    ax.set_ylabel('Degrees')
    ax.legend(loc='upper right')
    ax.grid(True, linestyle='--', alpha=0.7)
    ax.set_title(title)

plt.tight_layout(rect=[0, 0, 1, 0.95])

# --- GUI User Event ---
def on_key_press(event):
    global is_recording, is_marking, reset_to_zero
    if event.key == ' ':
        if not is_recording:
            is_recording = True
            print('🔴 [開始記錄] 現在開始將資料寫入 Excel 緩衝。再次按空白鍵停止並保存。')
        else:
            is_recording = False
            print('⏹️ [停止記錄] 正在保存已記錄資料到 Excel。')
            save_to_excel_now()
    elif event.key == '1':
        is_marking = not is_marking
        if is_marking:
            print('🟡 [標記開啟] 現在按 1 標記的資料會被置黃。再次按 1 關閉標記。')
        else:
            print('⚪ [標記關閉] 目前不再標記資料。')
    elif event.key.lower() == 'r':
        reset_to_zero = set(SENSOR_IDS)
        print('🔄 [重置] 目前姿勢將重新設為新 0° 基準')

fig.canvas.mpl_connect('key_press_event', on_key_press)

# --- Excel Save ---
def get_excel_filename():
    return os.path.join(LOG_DIRECTORY, f"CloseLook_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


def save_to_excel_now():
    if log_queue.empty():
        print('⚠️ 沒有任何記錄資料，無需保存 Excel。')
        return

    excel_filename = get_excel_filename()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('IMU_Data')
    headers = [
        'timestamp', 'sensor_id', 'q0', 'q1', 'q2', 'q3', 'roll', 'pitch', 'yaw', 'is_marked'
    ]
    ws.append(headers)
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    from openpyxl.cell import WriteOnlyCell

    try:
        while True:
            try:
                ts, sid, q0, q1, q2, q3, r, p, y, marked = log_queue.get_nowait()
                row = [ts, sid, round(q0, 6), round(q1, 6), round(q2, 6), round(q3, 6), round(r, 2), round(p, 2), round(y, 2), 'Y' if marked else '']
                cell_row = []
                for val in row:
                    cell = WriteOnlyCell(ws, value=val)
                    if marked:
                        cell.fill = yellow_fill
                    cell_row.append(cell)
                ws.append(cell_row)
            except queue.Empty:
                break
    finally:
        print(f'💾 正在保存 Excel 檔案到 {excel_filename} ...')
        wb.save(excel_filename)
        print(f'✅ Excel 檔案已保存成功: {excel_filename}')


# --- Math Core ---
def calculate_euler(q0, q1, q2, q3):
    try:
        pitch_val = max(-1.0, min(1.0, -2 * q1 * q3 + 2 * q0 * q2))
        pitch = math.degrees(math.asin(pitch_val))
        roll = math.degrees(math.atan2(2 * q2 * q3 + 2 * q0 * q1, -2 * q1 * q1 - 2 * q2 * q2 + 1))
        yaw = math.degrees(math.atan2(2 * q1 * q2 + 2 * q0 * q3, q0 * q0 + q1 * q1 - q2 * q2 - q3 * q3))
        return roll, pitch, yaw
    except ValueError:
        return 0.0, 0.0, 0.0


def update(frame):
    global start_time, is_calibrating, reset_to_zero
    updated_artists = []

    if ser:
        while ser.in_waiting > 0:
            try:
                line = ser.readline().decode('utf-8').strip()
                if not line:
                    continue
                parts = line.split(',')

                if len(parts) == 15:
                    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    curr_time_sec = time.time()
                    if start_time is None:
                        start_time = curr_time_sec

                    elapsed_time = curr_time_sec - start_time
                    if is_calibrating and elapsed_time > CALIBRATION_TIME_SEC:
                        is_calibrating = False
                        print('✅ Calibration finished! Baseline applied.')
                        for sid in SENSOR_IDS:
                            if calibration_data['roll'][sid]:
                                baseline_offsets['roll'][sid] = sum(calibration_data['roll'][sid]) / len(calibration_data['roll'][sid])
                                baseline_offsets['pitch'][sid] = sum(calibration_data['pitch'][sid]) / len(calibration_data['pitch'][sid])
                                yaws = np.radians(calibration_data['yaw'][sid])
                                mean_yaw_rad = np.arctan2(np.mean(np.sin(yaws)), np.mean(np.cos(yaws)))
                                baseline_offsets['yaw'][sid] = (math.degrees(mean_yaw_rad) + 360) % 360

                    for i in range(3):
                        idx = i * 5
                        sid = int(parts[idx])
                        if sid in SENSOR_IDS:
                            q0, q1, q2, q3 = map(float, parts[idx + 1: idx + 5])
                            roll, pitch, yaw = calculate_euler(q0, q1, q2, q3)
                            roll_out, pitch_out, yaw_out = roll, pitch, yaw

                            if is_calibrating:
                                calibration_data['roll'][sid].append(roll)
                                calibration_data['pitch'][sid].append(pitch)
                                calibration_data['yaw'][sid].append(yaw)
                                roll_out, pitch_out, yaw_out = 0.0, 0.0, 0.0
                            else:
                                if sid in reset_to_zero:
                                    delta_roll = roll - baseline_offsets['roll'][sid]
                                    delta_pitch = pitch - baseline_offsets['pitch'][sid]
                                    delta_yaw = yaw - baseline_offsets['yaw'][sid]
                                    baseline_offsets['roll'][sid] = roll
                                    baseline_offsets['pitch'][sid] = pitch
                                    baseline_offsets['yaw'][sid] = yaw
                                    for j in range(len(data['roll'][sid])):
                                        data['roll'][sid][j] -= delta_roll
                                        if data['roll'][sid][j] > 180:
                                            data['roll'][sid][j] -= 360
                                        elif data['roll'][sid][j] < -180:
                                            data['roll'][sid][j] += 360
                                        data['pitch'][sid][j] -= delta_pitch
                                        if data['pitch'][sid][j] > 180:
                                            data['pitch'][sid][j] -= 360
                                        elif data['pitch'][sid][j] < -180:
                                            data['pitch'][sid][j] += 360
                                        data['yaw'][sid][j] -= delta_yaw
                                        if data['yaw'][sid][j] > 180:
                                            data['yaw'][sid][j] -= 360
                                        elif data['yaw'][sid][j] < -180:
                                            data['yaw'][sid][j] += 360
                                    reset_to_zero.remove(sid)

                                roll_out = roll - baseline_offsets['roll'][sid]
                                pitch_out = pitch - baseline_offsets['pitch'][sid]
                                yaw_out = yaw - baseline_offsets['yaw'][sid]
                                if roll_out > 180:
                                    roll_out -= 360
                                elif roll_out < -180:
                                    roll_out += 360
                                if pitch_out > 180:
                                    pitch_out -= 360
                                elif pitch_out < -180:
                                    pitch_out += 360
                                if yaw_out > 180:
                                    yaw_out -= 360
                                elif yaw_out < -180:
                                    yaw_out += 360

                            data['roll'][sid].append(roll_out)
                            data['pitch'][sid].append(pitch_out)
                            data['yaw'][sid].append(yaw_out)

                            if is_recording:
                                log_queue.put((ts, sid, q0, q1, q2, q3, roll_out, pitch_out, yaw_out, is_marking))

            except (ValueError, IndexError, UnicodeDecodeError):
                pass

    for i, key in enumerate(data_keys):
        ax = axes[i]
        all_vals = []
        for sid in SENSOR_IDS:
            lines[key][sid].set_data(range(len(data[key][sid])), data[key][sid])
            updated_artists.append(lines[key][sid])
            if len(data[key][sid]) > 0:
                all_vals.extend(data[key][sid])

        if all_vals:
            avg_val = sum(all_vals) / len(all_vals)
            min_val = min(all_vals)
            max_val = max(all_vals)
            target_min = avg_val - 10.0
            target_max = avg_val + 10.0
            if min_val < target_min:
                target_min = min_val - 2.0
            if max_val > target_max:
                target_max = max_val + 2.0
            ax.set_ylim(target_min, target_max)

    fig.canvas.draw_idle()
    return updated_artists


# --- Run Application ---
if __name__ == '__main__':
    ani = animation.FuncAnimation(fig, update, interval=20, blit=False, cache_frame_data=False)
    try:
        plt.show()
    except KeyboardInterrupt:
        pass
    finally:
        running = False
        if not log_queue.empty():
            print('🔔 程式關閉時檢測到尚未保存的記錄，正在保存...')
            save_to_excel_now()
        if ser:
            ser.close()