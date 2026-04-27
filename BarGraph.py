import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

# ==========================================
# 1. FILE & ANGLE CONFIGURATION
# ==========================================
FILE_CONFIGS = [
    {'filename': 'Yaw0.xlsx',   'expected_yaw': 0.0},
    {'filename': 'Yaw30.xlsx',  'expected_yaw': 30.0},
    {'filename': 'Yaw45.xlsx',  'expected_yaw': 45.0},
    {'filename': 'Yaw60.xlsx',  'expected_yaw': 60.0},
    {'filename': 'Yaw90.xlsx',  'expected_yaw': 90.0},
    {'filename': 'Yaw-30.xlsx', 'expected_yaw': -30.0},
    {'filename': 'Yaw-45.xlsx', 'expected_yaw': -45.0},
    {'filename': 'Yaw-60.xlsx', 'expected_yaw': -60.0},
    {'filename': 'Yaw-90.xlsx', 'expected_yaw': -90.0},
]

OUTPUT_FILE = 'Batch_Validation_Output.xlsx'

# Highlight Configuration
HIGHLIGHT_TARGET = 'Yaw'  
HIGHLIGHT_COLOR = 'FFFFE0' # Light Yellow

ORDERED_SENSORS = ["21", "22", "23"]

# ==========================================
# 2. STATISTICAL ENGINE
# ==========================================
def calculate_yaw_stats(target_df, expected_val):
    """Calculates all scientific metrics for Yaw axis"""
    if target_df.empty:
        return {"Expected": expected_val, "Mean": "", "MAE": "", "RMSE": "", "Bias": "", "STD": "", "Lower_LOA": "", "Upper_LOA": ""}
    
    # 1. Raw Measured Average
    measured_mean = np.mean(target_df['yaw'])
    
    # 2. Error Calculations
    raw_err = target_df['yaw'] - expected_val
    abs_err = np.abs(raw_err)
    
    mae = np.mean(abs_err)
    rmse = np.sqrt(np.mean(raw_err**2))
    bias = np.mean(raw_err)
    std = np.std(raw_err)
    
    # 3. 95% Limits of Agreement
    lower_loa = bias - (1.96 * std)
    upper_loa = bias + (1.96 * std)
    
    return {
        "Expected": expected_val,
        "Mean": round(measured_mean, 3),
        "MAE": round(mae, 3),
        "RMSE": round(rmse, 3),
        "Bias": round(bias, 3),
        "STD": round(std, 3),
        "Lower_LOA": round(lower_loa, 3),
        "Upper_LOA": round(upper_loa, 3)
    }

# ==========================================
# 3. MAIN PROCESSING FUNCTION
# ==========================================
def process_batch():
    print("🚀 Starting 3-Axis Batch Processing...")
    
    # Dictionaries to hold the row data for each sheet
    results = {
        '21': [],
        '22': [],
        '23': []
    }
    
    for config in FILE_CONFIGS:
        fname = config['filename']
        print(f"📂 Processing: {fname}...")
        
        # --- Handle Missing Files ---
        if not os.path.exists(fname):
            print(f"   ⚠️ File not found. Creating blank row.")
            empty_row = {'File_Name': fname, 'Sample_Count': 0}
            empty_row.update({
                'Expected_Yaw': config['expected_yaw'],
                'Yaw_Mean': "",
                'Yaw_MAE': "", 'Yaw_RMSE': "", 'Yaw_Bias': "", 
                'Yaw_STD': "", 'Yaw_Lower_LOA': "", 'Yaw_Upper_LOA': ""
            })
            for key in results.keys():
                results[key].append(empty_row)
            continue
            
        # --- Read and Filter Data ---
        wb = load_workbook(fname)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        # Find column indices
        col_indices = {header: idx for idx, header in enumerate(headers)}
        
        if 'yaw' not in col_indices:
            print(f"   ❌ No 'yaw' column found in {fname}.")
            # Create empty data for missing files
            empty_df = SimpleDataFrame({'yaw': np.array([])})
            for sensor in ORDERED_SENSORS:
                sensor_df = SimpleDataFrame({'yaw': np.array([])})
                results[sensor].append(build_row(sensor_df))
            continue
            
        # Read and filter data
        paused_data = {header: [] for header in headers}
        for row in ws.iter_rows(min_row=2, values_only=True):
            is_paused = row[col_indices.get('is_paused', -1)] if 'is_paused' in col_indices else True
            if is_paused == True:
                for header, value in zip(headers, row):
                    paused_data[header].append(value)
        
        # Create simple data structure
        class SimpleDataFrame:
            def __init__(self, data_dict):
                self.data = data_dict
                first_key = list(data_dict.keys())[0] if data_dict else None
                self._len = len(data_dict[first_key]) if first_key else 0
            
            def __len__(self):
                return self._len
            
            @property
            def empty(self):
                return self._len == 0
            
            def __getitem__(self, key):
                if isinstance(key, str):
                    return np.array(self.data[key])
                else:
                    raise NotImplementedError
        
        paused_df = SimpleDataFrame(paused_data)
        
        # --- Helper Function to Build Row Data ---
        def build_row(sensor_data):
            row_data = {'File_Name': fname, 'Sample_Count': len(sensor_data)}
            
            # Process Yaw only
            expected = config['expected_yaw']
            stats = calculate_yaw_stats(sensor_data, expected)
            
            row_data.update({
                'Expected_Yaw': stats['Expected'],
                'Yaw_Mean': stats['Mean'],
                'Yaw_MAE': stats['MAE'],
                'Yaw_RMSE': stats['RMSE'],
                'Yaw_Bias': stats['Bias'],
                'Yaw_STD': stats['STD'],
                'Yaw_Lower_LOA': stats['Lower_LOA'],
                'Yaw_Upper_LOA': stats['Upper_LOA']
            })
            return row_data
        
        if paused_df.empty:
            print(f"   ❌ No paused data found in {fname}.")
            # Still add empty rows for this file
            for sensor in ORDERED_SENSORS:
                sensor_df = SimpleDataFrame({'yaw': np.array([])})
                results[sensor].append(build_row(sensor_df))
            continue

        # --- Process Individual Sensors ---
        for sensor in ORDERED_SENSORS:
            sensor_mask = np.array(paused_df.data['sensor_id']) == int(sensor)
            sensor_yaw_data = np.array(paused_df.data['yaw'])[sensor_mask]
            
            # Create sensor-specific DataFrame
            sensor_df = SimpleDataFrame({'yaw': sensor_yaw_data})
            results[sensor].append(build_row(sensor_df))

    # ==========================================
    # 4. EXCEL EXPORT & HIGHLIGHTING
    # ==========================================
    print(f"\n💾 Saving all tables to: {OUTPUT_FILE}")
    wb = Workbook()
    
    # Write each dictionary to a separate sheet
    for sheet_name, data_list in results.items():
        ws = wb.create_sheet(sheet_name)
        
        if not data_list:
            continue
            
        # Write headers
        headers = list(data_list[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, data_row in enumerate(data_list, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=data_row.get(header, ''))
        
        # Apply Highlighting if configured
        if HIGHLIGHT_TARGET:
            target_str = str(HIGHLIGHT_TARGET).lower()
            fill_style = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type='solid')
            
            # Check header row for the target string
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value and target_str in str(cell.value).lower():
                    # Highlight entire column
                    for row_idx in range(1, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill_style
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(OUTPUT_FILE)

    print(f"✨ Batch Processing Complete! Highlight target was set to '{HIGHLIGHT_TARGET}'.")
    
    # Generate validation plot
    generate_validation_plot(results)

# ==========================================
# 5. PLOT GENERATION
# ==========================================
def generate_validation_plot(results):
    """Generate IMU Yaw Accuracy Validation chart"""
    print("\n📊 Generating Yaw validation plot...")
    
    # Extract yaw angles from file names
    Yaw_angles = [0, 30, 45, 60, 90, -30, -45, -60, -90]
    
    # Prepare data for plotting - only individual sensors
    sensor_labels = ['21', '22', '23']
    sensor_colors = ['#1f77b4', '#2ca02c', '#ff7f0e']  # Blue, Green, Orange
    
    fig, ax = plt.subplots(figsize=(14, 8))
    
    x = np.arange(len(Yaw_angles))
    width = 0.25  # Wider bars since we have fewer sensors
    
    for idx, sensor in enumerate(sensor_labels):
        mae_values = []
        std_values = []
        
        for row_data in results[sensor]:
            # Extract Yaw_MAE and Yaw_STD from the row
            mae = row_data.get('Yaw_MAE', 0)
            std = row_data.get('Yaw_STD', 0)
            mae_values.append(mae if isinstance(mae, (int, float)) and not np.isnan(mae) else 0)
            std_values.append(std if isinstance(std, (int, float)) and not np.isnan(std) else 0)
        
        # Ensure we have the right number of values
        while len(mae_values) < len(Yaw_angles):
            mae_values.append(0)
            std_values.append(0)
        mae_values = mae_values[:len(Yaw_angles)]
        std_values = std_values[:len(Yaw_angles)]
        
        offset = (idx - 1) * width  # Center the three bars
        ax.bar(x + offset, mae_values, width, label=f'Sensor {sensor}',
               color=sensor_colors[idx], yerr=std_values, capsize=5, alpha=0.8)
    
    # Formatting
    ax.set_xlabel('Expected Yaw Angle (degrees)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Mean Absolute Error (degrees)', fontsize=12, fontweight='bold')
    ax.set_title('IMU Yaw Angle Accuracy Validation\nError bars show ±1 standard deviation', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels([f'{angle}°' for angle in Yaw_angles])
    ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1))
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig('Yaw_Validation_Chart.png', dpi=300, bbox_inches='tight')
    print("✅ Yaw validation chart saved as 'Yaw_Validation_Chart.png'")

if __name__ == "__main__":
    process_batch()